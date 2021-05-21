import json
import logging
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple

import cv2
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell import Cell as WsCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sklearn.cluster import DBSCAN

from table_extractor.bordered_service.models import InferenceTable, Page
from table_extractor.cascade_rcnn_service.inference import (
    CascadeRCNNInferenceService,
)
from excel_extractor.constants import HEADER_FILL
from excel_extractor.extractor import clean_xlsx_images
from table_extractor.headers.header_utils import HeaderChecker
from table_extractor.model.table import (
    BorderBox,
    Cell,
    CellLinked,
    StructuredTable,
    StructuredTableHeadered,
    TextField,
)
from table_extractor.pdf_service.pdf_to_image import convert_pdf_to_images
from table_extractor.pipeline.pipeline import page_to_dict

DEFAULT_WIDTH = 10

DEFAULT_HEIGHT = 15

HEADER_CHECKER = HeaderChecker()

NUMBER_FORMAT = re.compile("([0#,]+)(\.[0]+)*(%)*(E\+0+)*([^0E].*)*")

DATE_FORMAT_MAPPING = {
    "yyyy\\/mm\\/dd\\ hh:mm": "%Y/%m/%d %H:%M",
    "[h]:mm:ss": "%H:%M:%S",
    "d-mmm-yy": "%d-%b-%y",
    "[$-409]d\\-mmm\\-yyyy;@": "%d-%b-%Y",
}

LOGGER = logging.getLogger(__name__)

CASCADE_CONFIG_PATH = (
    Path(os.environ.get("CASCADE_CONFIG_PATH"))
    if os.environ.get("CASCADE_CONFIG_PATH")
    else Path(__file__).parent.parent.joinpath(
        "configs/cascadetabnet_config_5_cls_w18.py"
    )
)
CASCADE_MODEL_PATH = (
    Path(os.environ.get("CASCADE_MODEL_PATH"))
    if os.environ.get("CASCADE_MODEL_PATH")
    else Path(__file__).parent.parent.joinpath("models/epoch_20_w18.pth")
)


def save_document(document: Dict, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(path.absolute()), "w") as f:
        f.write(json.dumps(document, indent=4))


def create_copy_xlsx(orig_path: Path, out_path: Path, copy_name: str):
    out_path.mkdir(exist_ok=True, parents=True)
    exit_code = os.system(
        f"cp '{str(orig_path.absolute())}' '{str((out_path / copy_name).absolute())}'"
    )
    if exit_code != 0:
        raise ValueError("Can not copy file!")


def create_pdf(xlsx_path: Path, out_path: Path):
    out_path.mkdir(exist_ok=True, parents=True)
    exit_code = os.system(
        f"libreoffice --headless --convert-to pdf "
        f"--outdir '{str(out_path.absolute())}' '{str(xlsx_path.absolute())}'"
        f" -env:UserInstallation=file://{str(Path(__file__).parent.parent.joinpath('configs/').absolute())}"
    )
    if exit_code != 0:
        raise ValueError("Can not create PDF!")


def prepare_for_inference(xlsx_path, out_dir: Path) -> Path:
    create_copy_xlsx(xlsx_path, out_dir, "for_inference.xlsx")
    clean_xlsx_images(out_dir / "for_inference.xlsx")
    create_pdf(out_dir / "for_inference.xlsx", out_dir)
    return out_dir / "for_inference.pdf"


def softmax(array: Tuple[float]) -> List[float]:
    x = np.array(array)
    e_x = np.exp(x - np.max(x))
    return (e_x / e_x.sum()).tolist()


def _count_empty_cells(series: List[CellLinked]):
    return len([True for cell in series if cell.is_empty()])


def cell_in_inf_header(cell: CellLinked, inf_headers: List[Cell]) -> float:
    confidences = [0.0]
    for header in inf_headers:
        if cell.box_is_inside_another(header):
            confidences.append(header.confidence)
    return max(confidences)


def analyse(series: List[CellLinked], inf_headers: List[Cell]):
    # Check if series is header
    headers = []
    first_line = False
    for cell in series:
        inf_header_score = cell_in_inf_header(cell, inf_headers)
        header_score, cell_score = softmax(HEADER_CHECKER.get_cell_score(cell))
        header_score, cell_score = softmax(
            (header_score + inf_header_score, cell_score)
        )

        if header_score > cell_score:
            headers.append(cell)
        if cell.col == 0 and cell.row == 0:
            first_line = True
    if first_line:
        empty_cells_num = _count_empty_cells(series)
        return len(headers) > (len(series) - empty_cells_num) / 2
    # return len(headers) > (len(series) / 5) if len(series) > thresh else len(headers) > (len(series) / 2)
    return len(headers) > (len(series) / 2)


def create_header(
    series: List[List[CellLinked]], inf_headers: List[Cell], header_limit: int
):
    """
    Search for headers based on cells contents
    @param series: cols or rows of the table
    @param table:
    @param header_limit:
    """
    header_candidates = []
    last_header = None
    for idx, line in enumerate(series[:header_limit]):
        if analyse(line, inf_headers):
            header_candidates.append((idx, True, line))
            last_header = idx
        else:
            header_candidates.append((idx, False, line))

    if last_header is not None:
        header = [
            line
            for idx, is_header, line in header_candidates[: last_header + 1]
        ]
    else:
        header = []

    if len(header) > 0.75 * len(series):
        header = []

    return header


def get_headers_using_structured(
    table: StructuredTable, headers: List[Cell]
) -> StructuredTableHeadered:
    header_rows = create_header(table.rows, headers, 4)
    if not header_rows:
        header_rows = table.rows[:1]
    table_with_header = StructuredTableHeadered.from_structured_and_rows(
        table, header_rows
    )
    header_cols = create_header(table.cols, headers, 1)
    table_with_header.actualize_header_with_cols(header_cols)

    return table_with_header


def comp_table(worksheet: Worksheet,
               row_dim: List[float],
               col_dim: List[float],
               s_cell: Tuple[int, int],
               e_cell: Tuple[int, int],
               headers: List[Cell]):
    m_ranges = []
    for m_range in worksheet.merged_cells.ranges:
            m_ranges.append(m_range)
    s_row, s_col = s_cell
    e_row, e_col = e_cell

    cells = []
    m_range_included = []
    for row in range(s_row, e_row + 1):
        for col in range(s_col, e_col + 1):
            is_in_merged = False
            cur_m_range = None
            for m_range in m_ranges:
                if (row, col) in list(m_range.cells):
                    is_in_merged = True
                    cur_m_range = m_range
                    break
            skip = False
            if is_in_merged:
                for m_range in m_range_included:
                    if (row, col) in list(m_range.cells):
                        skip = True
                        break
            if skip:
                continue
            if is_in_merged and cur_m_range:
                m_range_included.append(cur_m_range)
                cells.append(CellLinked(
                    top_left_y=int(row_dim[cur_m_range.min_row - 1]),
                    top_left_x=int(col_dim[cur_m_range.min_col - 1]),
                    bottom_right_y=int(row_dim[cur_m_range.max_row]),
                    bottom_right_x=int(col_dim[cur_m_range.max_col]),
                    col=col - 1,
                    row=row - 1,
                    col_span=cur_m_range.max_col - cur_m_range.min_col + 1,
                    row_span=cur_m_range.max_row - cur_m_range.min_row + 1,
                    text_boxes=[
                        TextField(
                            bbox=BorderBox(
                                top_left_y=int(row_dim[cur_m_range.min_row - 1]),
                                top_left_x=int(col_dim[cur_m_range.min_col - 1]),
                                bottom_right_y=int(row_dim[cur_m_range.max_row]),
                                bottom_right_x=int(col_dim[cur_m_range.max_col]),
                            ),
                            text=extract_cell_value(
                                cur_m_range.start_cell
                            ),
                        )
                    ],
                ))
            else:
                cells.append(
                    CellLinked(
                        top_left_y=int(row_dim[row - 1]),
                        top_left_x=int(col_dim[col - 1]),
                        bottom_right_y=int(row_dim[row]),
                        bottom_right_x=int(col_dim[col]),
                        col=col - 1,
                        row=row - 1,
                        col_span=1,
                        row_span=1,
                        text_boxes=[
                            TextField(
                                bbox=BorderBox(
                                    top_left_y=int(row_dim[row - 1]),
                                    top_left_x=int(col_dim[col - 1]),
                                    bottom_right_y=int(row_dim[row]),
                                    bottom_right_x=int(col_dim[col]),
                                ),
                                text=extract_cell_value(
                                    worksheet.cell(row, col)
                                ),
                            )
                        ],
                    )
                )
    struct_table = StructuredTable(
        bbox=BorderBox(
            top_left_y=int(row_dim[s_row - 1]),
            top_left_x=int(col_dim[s_col - 1]),
            bottom_right_y=int(row_dim[e_row]),
            bottom_right_x=int(col_dim[e_col]),
        ),
        cells=cells,
    )
    struct_table_headered = get_headers_using_structured(struct_table, headers)
    head_cells = []
    for pack in struct_table_headered.header:
        head_cells.extend(pack)
    for cell in head_cells:
        col = cell.col + 1
        row = cell.row + 1
        col_span = cell.col_span
        row_span = cell.row_span
        for r in range(row, row + row_span):
            for c in range(col, col + col_span):
                worksheet.cell(r, c).fill = HEADER_FILL
    for cell in struct_table_headered.cells:
        col = cell.col + 1
        row = cell.row + 1
        col_span = cell.col_span
        row_span = cell.row_span
        for r in range(row, row + row_span):
            for c in range(col, col + col_span):
                worksheet.cell(r, c).fill = PatternFill(start_color="CC55BB", end_color="CC55BB", fill_type="solid")
    return struct_table_headered


def get_grid(worksheet: Worksheet, last_row: int, last_col: int):
    row_dim = [0]
    cur_row_dim = 0
    for row_id in range(1, last_row + 1):
        if worksheet.row_dimensions[row_id].height:
            cur_row_dim += worksheet.row_dimensions[row_id].height
        else:
            cur_row_dim += DEFAULT_HEIGHT
        row_dim.append(cur_row_dim)
    col_dim = [0]
    cur_col_dim = 0
    for col_id in range(1, last_col):
        if worksheet.column_dimensions[get_column_letter(col_id)].width:
            cur_col_dim += worksheet.column_dimensions[
                get_column_letter(col_id)
            ].width
        else:
            cur_col_dim += DEFAULT_WIDTH
        col_dim.append(cur_col_dim)
    return row_dim, col_dim


def clust_tables(worksheet: Worksheet, last_row: int, last_col: int):
    m_ranges_val = []
    for m_range in worksheet.merged_cells.ranges:
        if any(worksheet.cell(*cell).value for cell in m_range.cells):
            m_ranges_val.append(m_range)

    non_empty = []
    for i in range(1, last_row + 1):
        for j in range(1, last_col + 1):
            cell = worksheet.cell(i, j)
            if cell.value:
                non_empty.append([i, j])
            else:
                for m_range in m_ranges_val:
                    if (i, j) in list(m_range.cells):
                        non_empty.append([i, j])
                        break

    np_coords = np.array(non_empty)

    clust = DBSCAN(eps=1, min_samples=2).fit(np_coords)

    tables_cells = {}
    for cell_c, label in zip(non_empty, clust.labels_):
        if label in tables_cells:
            tables_cells[label].append(cell_c)
        else:
            tables_cells[label] = [cell_c]
    tables_proposals = []
    for table, cells in tables_cells.items():
        s_row = min([c[0] for c in cells])
        s_col = min([c[1] for c in cells])
        e_row = max([c[0] for c in cells])
        e_col = max([c[1] for c in cells])
        if e_row - s_row and e_col - s_col:
            tables_proposals.append((s_row, s_col, e_row, e_col))
    return tables_proposals


def extract_cell_value(ws_cell: WsCell):
    if not ws_cell or not ws_cell.value:
        return ""
    if ws_cell.data_type == "n":
        formats = NUMBER_FORMAT.findall(ws_cell.number_format)
        if formats:
            num_format = formats[0]
            measurement = num_format[4].replace("\\", "").replace('"', "")
            return f"{{:{',' if ',' in num_format[0] else ''}.{len(num_format[1]) - 1 if num_format[1] else 0}{'E' if num_format[3] else num_format[2] if num_format[2] else 'f'}}}{measurement}".format(
                ws_cell.value
            )
    if ws_cell.data_type == "d":
        form = DATE_FORMAT_MAPPING.get(ws_cell.number_format)
        if form:
            return ws_cell.value.strftime(form)
    return str(ws_cell.value)


def match_inf_res(xlsx_path: Path, images_dir: Path):
    LOGGER.info(
        "Initializing CascadeMaskRCNN with config: %s and model: %s",
        CASCADE_CONFIG_PATH,
        CASCADE_MODEL_PATH,
    )
    cascade_rcnn_detector = CascadeRCNNInferenceService(
        CASCADE_CONFIG_PATH, CASCADE_MODEL_PATH, True
    )
    pages = []
    workbook = load_workbook(str(xlsx_path.absolute()), data_only=True)
    for page_num, worksheet in enumerate(workbook.worksheets):
        row_fill = {}
        for row_id in range(1, worksheet.max_row + 1):
            row_fill[row_id] = False
            for col_id in range(1, worksheet.max_column + 1):
                if worksheet.cell(row_id, col_id).value:
                    row_fill[row_id] = True
                    break
        last_row = worksheet.max_row
        for row_id, not_empty in sorted(
            [(row_id, not_empty) for row_id, not_empty in row_fill.items()],
            reverse=True,
            key=lambda x: x[0],
        ):
            if not_empty:
                if last_row == worksheet.max_row:
                    last_row += 1
                break
            last_row = row_id

        col_fill = {}
        for col_id in range(1, worksheet.max_column + 1):
            col_fill[col_id] = False
            for row_id in range(1, worksheet.max_row + 1):
                if worksheet.cell(row_id, col_id).value:
                    col_fill[col_id] = True
                    break
        last_col = worksheet.max_column
        for col_id, not_empty in sorted(
            [(col_id, not_empty) for col_id, not_empty in col_fill.items()],
            reverse=True,
            key=lambda x: x[0],
        ):
            if not_empty:
                if last_col == worksheet.max_column:
                    last_col += 1
                break
            last_col = col_id

        height = 0
        for row_id in range(1, last_row):
            if worksheet.row_dimensions[row_id].height:
                height += worksheet.row_dimensions[row_id].height
            else:
                height += DEFAULT_HEIGHT
        width = 0
        for col_id in range(1, last_col):
            if worksheet.column_dimensions[get_column_letter(col_id)].width:
                width += worksheet.column_dimensions[
                    get_column_letter(col_id)
                ].width
            else:
                width += DEFAULT_WIDTH
        if height == 0 or width == 0:
            continue

        img = cv2.imread(str((images_dir / f"{page_num}.png").absolute()))
        img_shape = img.shape[:2]

        tables_proposals = clust_tables(worksheet, last_row, last_col)
        row_dim, col_dim = get_grid(worksheet, last_row, last_col)
        y_scale = img_shape[0] / height
        x_scale = img_shape[1] / width
        row_dim = [dim * y_scale for dim in row_dim]
        col_dim = [dim * x_scale for dim in col_dim]

        headers = []
        if last_row < 1000:
            _, headers = cascade_rcnn_detector.inference_image(
                images_dir / f"{page_num}.png", padding=200
            )
        tables = [comp_table(
            worksheet,
            row_dim,
            col_dim,
            (prop[0], prop[1]),
            (prop[2], prop[3]),
            headers
        ) for prop in tables_proposals]

        blocks = []
        blocks.extend(tables)
        prev_row_coord = 0
        for row_id in range(1, last_row):
            row_coord = prev_row_coord + (
                worksheet.row_dimensions[row_id].height
                if worksheet.row_dimensions[row_id].height
                else DEFAULT_HEIGHT
            )
            prev_col_coord = 0
            for col_id in range(1, last_col):
                col_coord = prev_col_coord + (
                    worksheet.column_dimensions[
                        get_column_letter(col_id)
                    ].width
                    if worksheet.column_dimensions[
                        get_column_letter(col_id)
                    ].width
                    else DEFAULT_WIDTH
                )
                if worksheet.cell(row_id, col_id).value and not any(
                    [
                        y1 <= row_id <= y2 and x1 <= col_id <= x2
                        for y1, x1, y2, x2 in tables_proposals
                    ]
                ):
                    text_field = TextField(
                        bbox=BorderBox(
                            top_left_x=prev_col_coord * x_scale,
                            top_left_y=prev_row_coord * y_scale,
                            bottom_right_x=col_coord * x_scale,
                            bottom_right_y=row_coord * y_scale,
                        ),
                        text=extract_cell_value(
                            worksheet.cell(row_id, col_id)
                        ),
                    )
                    blocks.append(text_field)
                prev_col_coord = col_coord
            prev_row_coord = row_coord

        pages.append(
            page_to_dict(
                Page(
                    page_num=page_num,
                    bbox=BorderBox(
                        top_left_x=0,
                        top_left_y=0,
                        bottom_right_x=img_shape[1],
                        bottom_right_y=img_shape[0],
                    ),
                    tables=blocks,
                )
            )
        )
    workbook.save(str(xlsx_path.absolute()))
    workbook.close()
    return pages


def crop_padding(
    inf_results: List[Tuple[List[InferenceTable], List[Cell]]], padding: int
):
    for inf_tables, headers in inf_results:
        for table in inf_tables:
            table.bbox.top_left_x = table.bbox.top_left_x - padding
            table.bbox.top_left_y = table.bbox.top_left_y - padding
            table.bbox.bottom_right_x = table.bbox.bottom_right_x - padding
            table.bbox.bottom_right_y = table.bbox.bottom_right_y - padding
            for cell in table.tags:
                cell.top_left_x = cell.top_left_x - padding
                cell.top_left_y = cell.top_left_y - padding
                cell.bottom_right_x = cell.bottom_right_x - padding
                cell.bottom_right_y = cell.bottom_right_y - padding
        for header in headers:
            header.top_left_x = header.top_left_x - padding
            header.top_left_y = header.top_left_y - padding
            header.bottom_right_x = header.bottom_right_x - padding
            header.bottom_right_y = header.bottom_right_y - padding


def run_excel_job(file: str, outpath: str):
    out_dir = Path(outpath) / Path(file).name
    pdf_path = prepare_for_inference(Path(file), out_dir)
    images_dir = convert_pdf_to_images(pdf_path, out_dir, True)

    create_copy_xlsx(Path(file), out_dir, "with_header.xlsx")

    pages = match_inf_res(out_dir / "with_header.xlsx", images_dir)

    document = {"doc_name": str(Path(file).name), "pages": pages}
    save_document(document, out_dir / "document.json")


if __name__ == "__main__":
    run_excel_job(
        [
            "/Users/matvei_vargasov/code/novatris/excel_extractor/base/samples/Examples.xlsx"
        ]
    )