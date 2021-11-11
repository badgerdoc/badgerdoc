import json
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple
import subprocess

import numpy as np
from openpyxl import load_workbook
from openpyxl.cell import Cell as WsCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sklearn.cluster import DBSCAN
from Levenshtein import distance as levenshtein_distance

from table_extractor.bordered_service.models import Page
from table_extractor.headers.header_utils import HeaderChecker
from table_extractor.model.table import (
    BorderBox,
    Cell,
    CellLinked,
    StructuredTable,
    StructuredTableHeadered,
    TextField,
)
from table_extractor.pipeline.pipeline import page_to_dict

DEFAULT_WIDTH = 10

DEFAULT_HEIGHT = 15

HEADER_CHECKER = HeaderChecker(
    cell_dictionary_path=Path(__file__).parent.parent.joinpath("language/cells_exc.json"),
    header_dictionary_path=Path(__file__).parent.parent.joinpath("language/headers_exc.json")
)

NUMBER_FORMAT = re.compile("([0#,]+)(\.[0]+)*(%)*(E\+0+)*([^0E].*)*")

DATE_FORMAT_MAPPING = {
    "yyyy\\/mm\\/dd\\ hh:mm": "%Y/%m/%d %H:%M",
    "[h]:mm:ss": "%H:%M:%S",
    "d-mmm-yy": "%d-%b-%y",
    "[$-409]d\\-mmm\\-yyyy;@": "%d-%b-%Y",
}

LOGGER = logging.getLogger(__name__)

HEADER_FILL = PatternFill(
    start_color="AA00CC", end_color="AA00CC", fill_type="solid"
)


@dataclass
class StructuredTableHeaderedEx(StructuredTableHeadered):

    @staticmethod
    def from_structured_and_rows(
            table: StructuredTable, header: List[List[CellLinked]]
    ):
        header_row_nums = set()
        for row in header:
            for cell in row:
                header_row_nums.add(cell.row)

        body_cells = [
            cell for cell in table.cells if cell.row not in header_row_nums
        ]

        return StructuredTableHeaderedEx(
            bbox=table.bbox, cells=body_cells, header=header
        )

    def actualize_header_with_cols(self, header_cols: List[List[CellLinked]]):
        header_col_nums = set()
        for col in header_cols:
            for cell in col:
                header_col_nums.add(cell.col)

        body_cells = [
            cell for cell in self.cells if cell.col not in header_col_nums
        ]

        self.header.extend(header_cols)
        self.cells = body_cells


def save_document(document: Dict, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(path.absolute()), "w") as f:
        f.write(json.dumps(document, indent=4))


def create_copy_xlsx(orig_path: Path, out_path: Path, copy_name: str):
    out_path.mkdir(exist_ok=True, parents=True)
    exit_code = os.system(
        f"cp '{str(orig_path.absolute())}' '{str((out_path / copy_name).absolute())}'"
    )
    if exit_code != 0:
        raise ValueError("Can not copy file!")


def convert_to_xlsx(xlsx_path: Path, out_path: Path):
    out_path.mkdir(exist_ok=True, parents=True)
    cmd = [
        f"{os.environ.get('LIBRE_RUN') if os.environ.get('LIBRE_RUN') else 'libreoffice'}",
        "--convert-to",
        "xlsx",
        f"{str(xlsx_path.absolute())}",
        "--outdir",
        f"{str(out_path.absolute())}",
        "--headless"
    ]
    LOGGER.info(f"Converting {xlsx_path}, command: {' '.join(cmd)}")
    process = subprocess.Popen(
        cmd,
        stdout=sys.stdout,
        stderr=sys.stderr
    )
    process.communicate()
    LOGGER.info(f"Exit code: {process.returncode}")
    if process.returncode != 0:
        raise ValueError("Can not create PDF!")


def softmax(array: Tuple[float]) -> List[float]:
    x = np.array(array)
    e_x = np.exp(x - np.max(x))
    return (e_x / e_x.sum()).tolist()


def _count_empty_cells(series: List[CellLinked]):
    return len([True for cell in series if cell.is_empty()])


def cell_in_inf_header(cell: CellLinked, inf_headers: List[Cell]) -> float:
    confidences = [0.0]
    for header in inf_headers:
        if cell.box_is_inside_another(header):
            confidences.append(header.confidence)
    return max(confidences)


def analyse(series: List[CellLinked]):
    # Check if series is header
    headers = []
    for cell in series:
        header_score, cell_score = softmax(HEADER_CHECKER.get_cell_score(cell))
        if header_score > cell_score:
            headers.append(cell)

    return len(headers) > (len(series) / 2)


def is_num(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


def create_header(
    series: List[List[CellLinked]], header_limit: int
):
    """
    Search for headers based on cells contents
    @param series: cols or rows of the table
    @param table:
    @param header_limit:
    """
    header_candidates = []
    last_header = None
    for idx, line in enumerate(series[:header_limit]):
        if analyse(line):
            header_candidates.append((idx, True, line))
            last_header = idx
        else:
            header_candidates.append((idx, False, line))

    if last_header is not None:
        header = [
            line
            for idx, is_header, line in header_candidates[: last_header + 1]
        ]
    else:
        header = []

    if len(header) > 0.65 * len(series):
        header = []

    if header_limit > 1:
        contents = [" ".join([cell.text_boxes[0].text for cell in line]) for line in series]
        l_dists = []
        last = contents[0]
        for line in contents[1:header_limit]:
            dist = levenshtein_distance(line, last)
            l_dists.append(dist)
            last = line
        index_max_lev = np.argmax(l_dists) + 1

        num_count = [len([val for val in line if is_num(val.text_boxes[0].text)]) for line in series]
        percentile_25 = np.percentile(num_count, 10)
        idis = []
        for idx, val in enumerate(num_count[:header_limit]):
            if val < percentile_25.astype(float):
                idis.append(idx)
        if not idis and not percentile_25:
            # n_dst = []
            # last = num_count[0]
            # for val in num_count[1:header_limit]:
            #     dst = val - last
            #     n_dst.append(dst)
            #     last = val
            # index_max_num = np.argmax(n_dst) + 1
            index_max_num = index_max_lev
        else:
            index_max_num = max(idis) + 1

        if not header:
            header = series[:index_max_num]

        if idis and percentile_25:
            header = series[:index_max_num]

    return header


def get_headers_using_structured(
    table: StructuredTable
) -> StructuredTableHeadered:
    header_rows = create_header(table.rows, 4)
    table_with_header = StructuredTableHeaderedEx.from_structured_and_rows(
        table, header_rows
    )
    header_cols = create_header(table.cols, 1)
    table_with_header.actualize_header_with_cols(header_cols)

    return table_with_header


def comp_table(worksheet: Worksheet,
               row_dim: List[float],
               col_dim: List[float],
               s_cell: Tuple[int, int],
               e_cell: Tuple[int, int]):
    m_ranges = []
    for m_range in worksheet.merged_cells.ranges:
            m_ranges.append(m_range)
    s_row, s_col = s_cell
    e_row, e_col = e_cell
    e_row = min(e_row, len(row_dim) - 1)
    e_col = min(e_col, len(col_dim) - 1)

    cells = []
    m_range_included = []
    for row in range(s_row, e_row + 1):
        for col in range(s_col, e_col + 1):
            is_in_merged = False
            cur_m_range = None
            for m_range in m_ranges:
                if (row, col) in list(m_range.cells):
                    is_in_merged = True
                    cur_m_range = m_range
                    break
            skip = False
            if is_in_merged:
                for m_range in m_range_included:
                    if (row, col) in list(m_range.cells):
                        skip = True
                        break
            if skip:
                continue
            if is_in_merged and cur_m_range:
                m_range_included.append(cur_m_range)
                cells.append(CellLinked(
                    top_left_y=int(row_dim[cur_m_range.min_row - 1]),
                    top_left_x=int(col_dim[cur_m_range.min_col - 1]),
                    bottom_right_y=int(row_dim[min(cur_m_range.max_row, len(row_dim) - 1)]),
                    bottom_right_x=int(col_dim[min(cur_m_range.max_col, len(col_dim) - 1)]),
                    col=col - 1,
                    row=row - 1,
                    col_span=cur_m_range.max_col - cur_m_range.min_col + 1,
                    row_span=cur_m_range.max_row - cur_m_range.min_row + 1,
                    text_boxes=[
                        TextField(
                            bbox=BorderBox(
                                top_left_y=int(row_dim[cur_m_range.min_row - 1]),
                                top_left_x=int(col_dim[cur_m_range.min_col - 1]),
                                bottom_right_y=int(row_dim[min(cur_m_range.max_row, len(row_dim) - 1)]),
                                bottom_right_x=int(col_dim[min(cur_m_range.max_col, len(col_dim) - 1)]),
                            ),
                            text=extract_cell_value(
                                cur_m_range.start_cell
                            ),
                        )
                    ],
                ))
            else:
                cells.append(
                    CellLinked(
                        top_left_y=int(row_dim[row - 1]),
                        top_left_x=int(col_dim[col - 1]),
                        bottom_right_y=int(row_dim[row]),
                        bottom_right_x=int(col_dim[col]),
                        col=col - 1,
                        row=row - 1,
                        col_span=1,
                        row_span=1,
                        text_boxes=[
                            TextField(
                                bbox=BorderBox(
                                    top_left_y=int(row_dim[row - 1]),
                                    top_left_x=int(col_dim[col - 1]),
                                    bottom_right_y=int(row_dim[row]),
                                    bottom_right_x=int(col_dim[col]),
                                ),
                                text=extract_cell_value(
                                    worksheet.cell(row, col)
                                ),
                            )
                        ],
                    )
                )
    struct_table = StructuredTable(
        bbox=BorderBox(
            top_left_y=int(row_dim[s_row - 1]),
            top_left_x=int(col_dim[s_col - 1]),
            bottom_right_y=int(row_dim[e_row]),
            bottom_right_x=int(col_dim[e_col]),
        ),
        cells=cells,
    )
    struct_table_headered = get_headers_using_structured(struct_table)
    if len(struct_table_headered.cells) + sum([len(h) for h in struct_table_headered.header]) > 3:
        head_cells = []
        for pack in struct_table_headered.header:
            head_cells.extend(pack)
        for cell in head_cells:
            col = cell.col + 1
            row = cell.row + 1
            col_span = cell.col_span
            row_span = cell.row_span
            for r in range(row, row + row_span):
                for c in range(col, col + col_span):
                    worksheet.cell(r, c).fill = HEADER_FILL
        for cell in struct_table_headered.cells:
            col = cell.col + 1
            row = cell.row + 1
            col_span = cell.col_span
            row_span = cell.row_span
            for r in range(row, row + row_span):
                for c in range(col, col + col_span):
                    worksheet.cell(r, c).fill = PatternFill(start_color="CC55BB", end_color="CC55BB", fill_type="solid")
    return struct_table_headered


def get_grid(worksheet: Worksheet, last_row: int, last_col: int):
    row_dim = [0]
    cur_row_dim = 0
    for row_id in range(1, last_row + 1):
        if worksheet.row_dimensions[row_id].height:
            cur_row_dim += worksheet.row_dimensions[row_id].height
        else:
            cur_row_dim += DEFAULT_HEIGHT
        row_dim.append(cur_row_dim)
    col_dim = [0]
    cur_col_dim = 0
    for col_id in range(1, last_col):
        if worksheet.column_dimensions[get_column_letter(col_id)].width:
            cur_col_dim += worksheet.column_dimensions[
                get_column_letter(col_id)
            ].width
        else:
            cur_col_dim += DEFAULT_WIDTH
        col_dim.append(cur_col_dim)
    return row_dim, col_dim


def clust_tables(worksheet: Worksheet, last_row: int, last_col: int):
    m_ranges_val = []
    for m_range in worksheet.merged_cells.ranges:
        if any(worksheet.cell(*cell).value for cell in m_range.cells):
            m_ranges_val.append(m_range)

    non_empty = []
    for i in range(1, last_row + 1):
        for j in range(1, last_col + 1):
            cell = worksheet.cell(i, j)
            if cell.value != '' and cell.value is not None:
                non_empty.append([i, j])
            else:
                for m_range in m_ranges_val:
                    if (i, j) in list(m_range.cells):
                        non_empty.append([i, j])
                        break

    np_coords = np.array(non_empty)

    clust = DBSCAN(eps=1.5, min_samples=2).fit(np_coords)

    tables_cells = {}
    for cell_c, label in zip(non_empty, clust.labels_):
        if label in tables_cells:
            tables_cells[label].append(cell_c)
        else:
            tables_cells[label] = [cell_c]
    tables_proposals = []
    for table, cells in tables_cells.items():
        s_row = min([c[0] for c in cells])
        s_col = min([c[1] for c in cells])
        e_row = max([c[0] for c in cells])
        e_col = max([c[1] for c in cells])
        if e_row - s_row and e_col - s_col:
            tables_proposals.append((s_row, s_col, e_row, e_col))

    tbl_prop = tables_proposals.copy()
    tbls = []
    while tbl_prop:
        tbl = tbl_prop.pop()
        table_int = []
        s_row, s_col, e_row, e_col = tbl
        for t_prop in tbl_prop:
            p_s_row, p_s_col, p_e_row, p_e_col = t_prop
            x_left = max(s_col, p_s_col)
            y_top = max(s_row, p_s_row)
            x_right = min(e_col, p_e_col)
            y_bottom = min(e_row, p_e_row)
            if x_right < x_left or y_bottom < y_top:
                table_int.append((p_s_row, p_s_col, p_e_row, p_e_col))
            else:
                s_row = min(s_row, p_s_row)
                s_col = min(s_col, p_s_col)
                e_row = max(e_row, p_e_row)
                e_col = max(e_col, p_e_col)
        tbls.append((s_row, s_col, e_row, e_col))
        tbl_prop = table_int
    return tbls


def extract_cell_value(ws_cell: WsCell):
    if not ws_cell or not ws_cell.value:
        return ""
    if ws_cell.data_type == "n":
        formats = NUMBER_FORMAT.findall(ws_cell.number_format)
        if formats:
            num_format = formats[0]
            measurement = num_format[4].replace("\\", "").replace('"', "")
            return f"{{:{',' if ',' in num_format[0] else ''}.{len(num_format[1]) - 1 if num_format[1] else 0}{'E' if num_format[3] else num_format[2] if num_format[2] else 'f'}}}{measurement}".format(
                ws_cell.value
            )
    if ws_cell.data_type == "d":
        form = DATE_FORMAT_MAPPING.get(ws_cell.number_format)
        if form:
            return ws_cell.value.strftime(form)
    return str(ws_cell.value)


def match_inf_res(xlsx_path: Path,):
    pages = []
    workbook = load_workbook(str(xlsx_path.absolute()), data_only=True)
    for page_num, worksheet in enumerate(workbook.worksheets):
        LOGGER.info(f"Processing sheet: {worksheet.title}")
        if worksheet.sheet_state != 'visible':
            worksheet.sheet_state = 'visible'
        row_fill = {}
        for row_id in range(1, worksheet.max_row + 1):
            row_fill[row_id] = False
            for col_id in range(1, worksheet.max_column + 1):
                if worksheet.cell(row_id, col_id).value:
                    row_fill[row_id] = True
                    break
        last_row = worksheet.max_row
        for row_id, not_empty in sorted(
            [(row_id, not_empty) for row_id, not_empty in row_fill.items()],
            reverse=True,
            key=lambda x: x[0],
        ):
            if not_empty:
                if last_row == worksheet.max_row:
                    last_row += 1
                break
            last_row = row_id

        col_fill = {}
        for col_id in range(1, worksheet.max_column + 1):
            col_fill[col_id] = False
            for row_id in range(1, worksheet.max_row + 1):
                if worksheet.cell(row_id, col_id).value:
                    col_fill[col_id] = True
                    break
        last_col = worksheet.max_column
        for col_id, not_empty in sorted(
            [(col_id, not_empty) for col_id, not_empty in col_fill.items()],
            reverse=True,
            key=lambda x: x[0],
        ):
            if not_empty:
                if last_col == worksheet.max_column:
                    last_col += 1
                break
            last_col = col_id

        height = 0
        for row_id in range(1, last_row):
            if worksheet.row_dimensions[row_id].height:
                height += worksheet.row_dimensions[row_id].height
            else:
                height += DEFAULT_HEIGHT
        width = 0
        for col_id in range(1, last_col):
            if worksheet.column_dimensions[get_column_letter(col_id)].width:
                width += worksheet.column_dimensions[
                    get_column_letter(col_id)
                ].width
            else:
                width += DEFAULT_WIDTH
        if height == 0 or width == 0:
            continue
        LOGGER.info(f"Max row: {last_row}, max col: {last_col}")

        tables_proposals = clust_tables(worksheet, last_row, last_col)
        LOGGER.info(f"Tables count: {len(tables_proposals)}")

        row_dim, col_dim = get_grid(worksheet, last_row, last_col)

        tables = [comp_table(
            worksheet,
            row_dim,
            col_dim,
            (prop[0], prop[1]),
            (prop[2], prop[3]),
        ) for prop in tables_proposals]

        tables = [table for table in tables if len(table.cells) + sum([len(h) for h in table.header]) > 3]

        blocks = []
        blocks.extend(tables)
        prev_row_coord = 0
        for row_id in range(1, last_row):
            row_coord = prev_row_coord + (
                worksheet.row_dimensions[row_id].height
                if worksheet.row_dimensions[row_id].height
                else DEFAULT_HEIGHT
            )
            prev_col_coord = 0
            for col_id in range(1, last_col):
                col_coord = prev_col_coord + (
                    worksheet.column_dimensions[
                        get_column_letter(col_id)
                    ].width
                    if worksheet.column_dimensions[
                        get_column_letter(col_id)
                    ].width
                    else DEFAULT_WIDTH
                )
                if worksheet.cell(row_id, col_id).value and not any(
                    [
                        y1 <= row_id <= y2 and x1 <= col_id <= x2
                        for y1, x1, y2, x2 in tables_proposals
                    ]
                ):
                    text_field = TextField(
                        bbox=BorderBox(
                            top_left_x=prev_col_coord,
                            top_left_y=prev_row_coord,
                            bottom_right_x=col_coord,
                            bottom_right_y=row_coord,
                        ),
                        text=extract_cell_value(
                            worksheet.cell(row_id, col_id)
                        ),
                    )
                    blocks.append(text_field)
                prev_col_coord = col_coord
            prev_row_coord = row_coord

        pages.append(
            page_to_dict(
                Page(
                    page_num=page_num,
                    bbox=BorderBox(
                        top_left_x=0,
                        top_left_y=0,
                        bottom_right_x=col_dim[-1],
                        bottom_right_y=row_dim[-1],
                    ),
                    tables=blocks,
                )
            )
        )
    workbook.save(str(xlsx_path.absolute()))
    workbook.close()
    return pages


def clean_xlsx_images(xlsx_path: Path):
    workbook = load_workbook(str(xlsx_path.absolute()))
    for worksheet in workbook.worksheets:
        worksheet._images = []
        worksheet._charts = []
    workbook.save(str(xlsx_path.absolute()))
    workbook.close()


def run_excel_job(file: str, outpath: str):
    out_dir = Path(outpath) / Path(file).name
    file = Path(file)
    if file.name.endswith(".xls"):
        convert_to_xlsx(file, out_dir)
        file = out_dir / file.name.replace(".xls", ".xlsx")

    if file.name.endswith(".csv"):
        convert_to_xlsx(file, out_dir)
        file = out_dir / file.name.replace(".csv", ".xlsx")

    create_copy_xlsx(file, out_dir, "with_header.xlsx")

    clean_xlsx_images(out_dir / "with_header.xlsx")

    pages = match_inf_res(out_dir / "with_header.xlsx")

    document = {"doc_name": str(file.name), "pages": pages}
    save_document(document, out_dir / "document.json")
