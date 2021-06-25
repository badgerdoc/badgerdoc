import json
import logging
import sys
from pathlib import Path

import click
from openpyxl import load_workbook
from openpyxl.cell import Cell as WsCell
from openpyxl.utils import coordinate_to_tuple

from table_extractor.headers.concordance_pandas import process_words

MARKUP_HEADER_COLOR = 'FF0066CC'
LOGGER = logging.getLogger(__name__)

LOGGING_FORMAT = "[%(asctime)s] - [%(name)s] - [%(levelname)s] - %(message)s"


def configure_logging():
    formatter = logging.Formatter(LOGGING_FORMAT)
    console_handler = logging.StreamHandler(stream=sys.stdout)
    console_handler.setFormatter(formatter)
    file_handler = logging.FileHandler(
        str(
            Path(__file__)
                .parent.joinpath("header_extractor.log")
                .absolute()
        )
    )
    file_handler.setFormatter(formatter)
    logging.root.setLevel(logging.DEBUG)
    logging.root.addHandler(console_handler)
    logging.root.addHandler(file_handler)


def resolve_header_color(
        header_color: str,
        header_excel: str,
        header_sheet: str,
        header_cell: str
):
    if any([not item for item in (header_excel, header_sheet, header_cell)]):
        return header_color
    LOGGER.info(f"Provided excel source to get header color: {header_excel}")
    workbook = load_workbook(header_excel, data_only=True)
    LOGGER.info(f"Provided sheet name  to get header color: {header_sheet}")
    sheet = workbook.get_sheet_by_name(header_sheet)
    color = sheet.cell(*coordinate_to_tuple(header_cell)).fill.bgColor.value
    workbook.close()
    return color


def extract_cell_value(ws_cell: WsCell):
    if not ws_cell or not ws_cell.value:
        return ""
    return str(ws_cell.value)


def extract_values(excel_markup_dir, header_color):
    color_cells = {}
    for file in Path(excel_markup_dir).glob("**/*.xlsx"):
        LOGGER.info(f"Processing file: {str(file.absolute())}")
        workbook = load_workbook(str(file.absolute()), data_only=True)
        for sheet in workbook.worksheets:
            for row_id in range(1, sheet.max_row + 1):
                for col_id in range(1, sheet.max_column + 1):
                    c = sheet.cell(row_id, col_id).fill.bgColor.value
                    if c in color_cells:
                        color_cells[c].append(sheet.cell(row_id, col_id))
                    else:
                        color_cells[c] = [sheet.cell(row_id, col_id)]
        workbook.close()
    headers = []
    non_headers = []
    for color, cells in color_cells.items():
        cell_values = [extract_cell_value(cell) for cell in cells if cell.value]
        if color == header_color:
            headers.extend(cell_values)
        else:
            non_headers.extend(cell_values)
    return headers, non_headers


@click.command()
@click.argument("excel_markup_dir")
@click.argument("dict_folder")
@click.option("--header_color", default=MARKUP_HEADER_COLOR)
@click.option("--header_excel")
@click.option("--header_sheet")
@click.option("--header_cell")
def extract_header(excel_markup_dir: str,
                   dict_folder: str,
                   header_color: str,
                   header_excel: str,
                   header_sheet: str,
                   header_cell: str
                   ):
    configure_logging()
    header_color = resolve_header_color(
        header_color,
        header_excel,
        header_sheet,
        header_cell
    )
    LOGGER.info(f"Resolved header color: {header_color}")
    headers, non_headers = extract_values(excel_markup_dir, header_color)
    headers_dict = process_words('headers', headers)
    cells_dict = process_words('cells', non_headers)
    Path(dict_folder).mkdir(exist_ok=True, parents=True)
    with open(f'{dict_folder}/headers.json', 'w') as f:
        f.write(json.dumps(headers_dict))
    with open(f'{dict_folder}/cells.json', 'w') as f:
        f.write(json.dumps(cells_dict))


if __name__ == '__main__':
    extract_header()
