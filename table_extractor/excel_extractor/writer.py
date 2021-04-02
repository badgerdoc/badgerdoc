from typing import Dict

from openpyxl import Workbook

from table_extractor.bordered_service.models import Page
from table_extractor.excel_extractor.constants import HEADER_FILL, HEADER_FONT
from table_extractor.excel_extractor.converter import (
    get_headers_using_structured,
)
from table_extractor.model.table import BorderBox
from table_extractor.pipeline.pipeline import page_to_dict


class BaseWriter:
    """
    Base class for writers
    """

    def __init__(self, data: Dict[str, list], outpath: str):
        self.data = data
        self.outpath = outpath
        self._tables_with_headers = None

    def write(self):
        raise NotImplemented

    @property
    def converted_data(self) -> dict:
        raise NotImplemented


class ExcelWriter(BaseWriter):
    """
    Generate excel from structured table
    """

    wb = Workbook()

    machine_learning_used = False

    def write(self):
        pages = []
        for i, (sheet, tables) in enumerate(self.tables_with_headers.items()):
            if not i:
                ws = self.wb.active
            else:
                ws = self.wb.create_sheet(sheet)

            for table in tables:
                for header_cells in table.header:
                    for cell in header_cells:
                        added_cell = ws.cell(
                            row=cell.row,
                            column=cell.col,
                            value=cell.text_boxes[0].text,
                        )
                        added_cell.fill = HEADER_FILL
                        added_cell.font = HEADER_FONT

                for cell in table.cells:
                    ws.cell(
                        row=cell.row,
                        column=cell.col,
                        value=cell.text_boxes[0].text,
                    )

            pages.append(
                page_to_dict(
                    Page(
                        page_num=i,
                        bbox=BorderBox(
                            top_left_x=0,
                            top_left_y=0,
                            bottom_right_x=max(
                                [table.bbox.bottom_right_x for table in tables]
                            ),
                            bottom_right_y=max(
                                [table.bbox.bottom_right_y for table in tables]
                            ),
                        ),
                        tables=tables,
                    )
                )
            )

        self.wb.save(self.outpath)
        return pages

    @property
    def tables_with_headers(self) -> dict:
        if not self._tables_with_headers:
            self._tables_with_headers = {}
            for sheet, tables in self.data.items():
                self._tables_with_headers[sheet] = [
                    get_headers_using_structured(table) for table in tables
                ]
        return self._tables_with_headers
