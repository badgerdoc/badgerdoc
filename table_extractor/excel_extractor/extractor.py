import sys

from typing import Union, Generator, List
from enum import Enum
from pathlib import Path
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell


class Output(Enum):
    PDF = 'pdf'
    EXCEL = 'excel'
    HTML = 'html'
    JSON = 'json'
    PNG = 'png'


@dataclass
class ExcelExtractor:
    """
    Render excel table to png file
    """

    tables = {}
    ws_tables = []
    coordinates = {}

    coordinates_offset_x: int = 1
    coordinates_offset_y: int = 1

    def extract(self, file: Union[str, Path]) -> dict:
        """
        Exctract information and save it in given ouput format
        """
        wb = load_workbook(file)

        for sheet in wb:
            self.parse_sheet(sheet)
            self.tables[sheet.title] = self.ws_tables
            self.ws_tables = []

        return self.tables

    def _check_interception(self, attr, table, cell) -> bool:
        index = int(attr == 'row')
        return table['end'][index] >= getattr(cell, attr) >= table['start'][index]

    def is_merged(self, cell) -> bool:
        return isinstance(cell, MergedCell)

    def iter_subtable_rows(self, sheet, first_cell: Cell, cell) -> Generator:
        max_col = cell.column if sheet.max_column == cell.column else cell.column - 1
        for row in sheet.iter_rows(min_row=first_cell.row, min_col=first_cell.column, max_col=max_col):
            yield row

    def in_table(self, cell: Cell) -> bool:
        for table in self.ws_tables:
            if self._check_interception('row', table, cell) and self._check_interception('column', table, cell):
                return True
        return False

    def get_end_cell(self, sheet: Worksheet, first_cell: Cell, current_cell: Cell) -> Cell:
        """
        Currently not used
        """
        row_number = first_cell.row
        for row in self.iter_subtable_rows(sheet, first_cell):
            row_number += 1

        column = current_cell.column if current_cell.column == sheet.max_column else current_cell.column - 1
        row_number -= 1
        return sheet.cell(row=row_number, column=column)

    def fill_cells(self, sheet: Worksheet, first_cell: Cell, last_cell: Cell) -> dict:
        cells = {}
        for row in list(self.iter_subtable_rows(sheet, first_cell, last_cell)):
            for cell in row:
                colspan = 0
                rowspan = 0

                if self.is_merged(cell):
                    column_letter = cell.coordinate[0]
                else:
                    column_letter = cell.column_letter
                    for merged_range in sheet.merged_cells.ranges:
                        if merged_range.start_cell == cell:
                            colspan = merged_range.size['columns']
                            rowspan = merged_range.size['rows']

                cells[(cell.column, cell.row)] = (
                    self.coordinates[cell.column, cell.row],
                    colspan,
                    rowspan,
                    str(cell.value) if cell.value else '',
                    cell,
                )
        return cells

    def get_table(self, sheet: Worksheet, cell: Cell, table_started: Cell) -> dict:
        """
        Returns dict with data table
        """
        table = {}
        table['start'] = (table_started.column, table_started.row)
        table['cells'] = self.fill_cells(sheet, table_started, cell)
        if not table['cells']:
            return None
        table['end'] = max(table['cells'].keys())
        table['dimensions'] = (
            self.coordinates[table['start']],
            self.coordinates[table['end']]
        )
        return table

    def finish_table(self, sheet, cell, table_started):
        table = self.get_table(sheet, cell, table_started)
        if table:
            self.ws_tables.append(table)

    def fill_coordinates(self, sheet):
        y_counter = 0
        for row in sheet.iter_rows():
            x_counter = 0
            max_height_cell = 0
            for cell in row:
                cell_width = sheet.column_dimensions[cell.coordinate[0]].width
                cell_height = sheet.row_dimensions[cell.row].height if sheet.row_dimensions[cell.row].height else 5.0
                if cell_height > max_height_cell:
                    max_height_cell = cell_height
                self.coordinates[(cell.column, cell.row)] = {
                    'top_left': (x_counter + self.coordinates_offset_x, y_counter + self.coordinates_offset_y),
                    'top_right': (
                        x_counter + cell_width - self.coordinates_offset_x, y_counter + self.coordinates_offset_y),
                    'bottom_left': (
                        x_counter + self.coordinates_offset_x, y_counter + cell_height - self.coordinates_offset_y),
                    'bottom_right': (x_counter + cell_width - self.coordinates_offset_x,
                                     y_counter + cell_height - self.coordinates_offset_y),
                    'width': cell_width,
                    'height': cell_height
                }
                x_counter += cell_width
            y_counter += max_height_cell

    def parse_sheet(self, sheet: Worksheet) -> list:
        tables = []
        self.fill_coordinates(sheet)
        self.finish_table(sheet, sheet[sheet.max_row][sheet.max_column-1], sheet[1][0])

        return tables


if __name__ == "__main__":
    ext = ExcelExtractor()
    #
