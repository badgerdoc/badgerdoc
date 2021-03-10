import json
import logging
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple

import cv2
from openpyxl.cell import Cell as WsCell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from table_extractor.bordered_service.models import InferenceTable, Page
from table_extractor.cascade_rcnn_service.inference import CascadeRCNNInferenceService
from table_extractor.excel_extractor.constants import HEADER_FILL
from table_extractor.excel_extractor.extractor import ExcelExtractor, clean_xlsx_images
from table_extractor.excel_extractor.converter import excel_to_structured, get_header_using_styles, \
    get_headers_using_structured
from table_extractor.excel_extractor.writer import ExcelWriter
from table_extractor.pdf_service.pdf_to_image import convert_pdf_to_images
from table_extractor.model.table import (
    StructuredTable,
    StructuredTableHeadered,
    BorderBox,
    CellLinked,
    TextField, Cell
)
from table_extractor.headers.header_utils import HeaderChecker
import numpy as np

from table_extractor.pipeline.pipeline import page_to_dict

DEFAULT_WIDTH = 10

DEFAULT_HEIGHT = 15

HEADER_CHECKER = HeaderChecker()

NUMBER_FORMAT = re.compile('([0#,]+)(\.[0]+)*(%)*(E\+0+)*([^0E].*)*')

DATE_FORMAT_MAPPING = {
    'yyyy\\/mm\\/dd\\ hh:mm': '%Y/%m/%d %H:%M',
    '[h]:mm:ss': '%H:%M:%S',
    'd-mmm-yy': '%d-%b-%y',
    '[$-409]d\\-mmm\\-yyyy;@': '%d-%b-%Y'
}

LOGGER = logging.getLogger(__name__)

CASCADE_CONFIG_PATH = Path(os.environ.get("CASCADE_CONFIG_PATH")) if os.environ.get("CASCADE_CONFIG_PATH") \
    else Path(__file__).parent.parent.parent.joinpath("configs/cascadetabnet_config_5_cls.py")
CASCADE_MODEL_PATH = Path(os.environ.get("CASCADE_MODEL_PATH")) if os.environ.get("CASCADE_MODEL_PATH") \
    else Path(__file__).parent.parent.parent.joinpath("models/epoch_10_scale_4.pth")


def save_document(document: Dict, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(path.absolute()), 'w') as f:
        f.write(json.dumps(document, indent=4))


def create_copy_xlsx(orig_path: Path, out_path: Path, copy_name: str):
    out_path.mkdir(exist_ok=True, parents=True)
    exit_code = os.system(f"cp '{str(orig_path.absolute())}' '{str((out_path / copy_name).absolute())}'")
    if exit_code != 0:
        raise ValueError("Can not copy file!")


def create_pdf(xlsx_path: Path, out_path: Path):
    out_path.mkdir(exist_ok=True, parents=True)
    exit_code = os.system(f"libreoffice --headless --convert-to pdf "
                          f"--outdir '{str(out_path.absolute())}' '{str(xlsx_path.absolute())}'"
                          f" -env:UserInstallation=file://{str(Path(__file__).parent.parent.parent.joinpath('configs/').absolute())}")
    if exit_code != 0:
        raise ValueError("Can not create PDF!")


def prepare_for_inference(xlsx_path, out_dir: Path) -> Path:
    create_copy_xlsx(xlsx_path, out_dir, 'for_inference.xlsx')
    clean_xlsx_images(out_dir / 'for_inference.xlsx')
    create_pdf(out_dir / 'for_inference.xlsx', out_dir)
    return out_dir / 'for_inference.pdf'


def softmax(array: Tuple[float]) -> List[float]:
    x = np.array(array)
    e_x = np.exp(x - np.max(x))
    return (e_x / e_x.sum()).tolist()


def _count_empty_cells(series: List[CellLinked]):
    return len([True for cell in series if cell.is_empty()])


def cell_in_inf_header(cell: CellLinked, inf_headers: List[Cell]) -> float:
    confidences = [0.]
    for header in inf_headers:
        if cell.box_is_inside_another(header):
            confidences.append(header.confidence)
    return max(confidences)


def analyse(series: List[CellLinked], inf_headers: List[Cell]):
    # Check if series is header
    headers = []
    first_line = False
    for cell in series:
        inf_header_score = cell_in_inf_header(cell, inf_headers)
        header_score, cell_score = softmax(HEADER_CHECKER.get_cell_score(cell))
        header_score, cell_score = softmax((header_score + inf_header_score, cell_score))

        if header_score > cell_score:
            headers.append(cell)
        if cell.col == 0 and cell.row == 0:
            first_line = True
    if first_line:
        empty_cells_num = _count_empty_cells(series)
        return len(headers) > (len(series) - empty_cells_num) / 2
    # return len(headers) > (len(series) / 5) if len(series) > thresh else len(headers) > (len(series) / 2)
    return len(headers) > (len(series) / 2)


def create_header(series: List[List[CellLinked]], inf_headers: List[Cell], header_limit: int):
    """
    Search for headers based on cells contents
    @param series: cols or rows of the table
    @param table:
    @param header_limit:
    """
    header_candidates = []
    last_header = None
    for idx, line in enumerate(series[:header_limit]):
        if analyse(line, inf_headers):
            header_candidates.append((idx, True, line))
            last_header = idx
        else:
            header_candidates.append((idx, False, line))

    if last_header is not None:
        header = [line for idx, is_header, line in header_candidates[:last_header + 1]]
    else:
        header = []

    if len(header) > 0.75 * len(series):
        header = []

    return header


def get_headers_using_structured(table: StructuredTable, headers: List[Cell]) -> StructuredTableHeadered:
    header_rows = create_header(table.rows, headers, 4)
    if not header_rows:
        header_rows = table.rows[:1]
    table_with_header = StructuredTableHeadered.from_structured_and_rows(table, header_rows)
    header_cols = create_header(table.cols, headers, 1)
    table_with_header.actualize_header_with_cols(header_cols)

    return table_with_header


def compute_tables(worksheet: Worksheet, last_row: int, last_col: int):
    cells = []
    height = 0
    width = 0
    for row in range(1, last_row):
        width = 0
        for col in range(1, last_col):
            cells.append(
                CellLinked(
                    top_left_y=height,
                    top_left_x=width,
                    bottom_right_y=height + DEFAULT_HEIGHT,
                    bottom_right_x=width + DEFAULT_WIDTH,
                    col=col - 1,
                    row=row - 1,
                    col_span=1,
                    row_span=1,
                    text_boxes=[
                        TextField(
                            bbox=BorderBox(
                                top_left_y=height,
                                top_left_x=width,
                                bottom_right_y=height + DEFAULT_HEIGHT,
                                bottom_right_x=width + DEFAULT_WIDTH,
                            ),
                            text=str(worksheet.cell(row, col).value)
                            if worksheet.cell(row, col).value else ''
                        )
                    ]
                )
            )
            width = width + DEFAULT_WIDTH
        height = height + DEFAULT_HEIGHT
    struct_table = StructuredTable(
        bbox=BorderBox(
            top_left_y=0,
            top_left_x=0,
            bottom_right_x=width,
            bottom_right_y=height,
        ),
        cells=cells
    )
    struct_table_headered = get_headers_using_structured(struct_table, [])
    head_cells = []
    for pack in struct_table_headered.header:
        head_cells.extend(pack)
    for cell in head_cells:
        worksheet.cell(cell.row + 1, cell.col + 1).fill = HEADER_FILL
    return struct_table_headered


def extract_cell_value(ws_cell: WsCell):
    if not ws_cell or not ws_cell.value:
        return ''
    if ws_cell.data_type == 'n':
        formats = NUMBER_FORMAT.findall(ws_cell.number_format)
        if formats:
            num_format = formats[0]
            measurement = num_format[4].replace('\\', '').replace('"', '')
            return f"{{:{',' if ',' in num_format[0] else ''}.{len(num_format[1]) - 1 if num_format[1] else 0}{'E' if num_format[3] else num_format[2] if num_format[2] else 'f'}}}{measurement}".format(
                ws_cell.value)
    if ws_cell.data_type == 'd':
        form = DATE_FORMAT_MAPPING.get(ws_cell.number_format)
        if form:
            return ws_cell.value.strftime(form)
    return str(ws_cell.value)


def match_inf_res(xlsx_path: Path,
                  images_dir: Path):
    LOGGER.info("Initializing CascadeMaskRCNN with config: %s and model: %s", CASCADE_CONFIG_PATH, CASCADE_MODEL_PATH)
    cascade_rcnn_detector = CascadeRCNNInferenceService(CASCADE_CONFIG_PATH, CASCADE_MODEL_PATH, True)
    pages = []
    workbook = load_workbook(str(xlsx_path.absolute()), data_only=True)
    for page_num, worksheet in enumerate(workbook.worksheets):
        row_fill = {}
        for row_id in range(1, worksheet.max_row + 1):
            row_fill[row_id] = False
            for col_id in range(1, worksheet.max_column + 1):
                if worksheet.cell(row_id, col_id).value:
                    row_fill[row_id] = True
                    break
        last_row = worksheet.max_row
        for row_id, not_empty in sorted([(row_id, not_empty) for row_id, not_empty in row_fill.items()], reverse=True,
                                        key=lambda x: x[0]):
            if not_empty:
                if last_row == worksheet.max_row:
                    last_row += 1
                break
            last_row = row_id

        col_fill = {}
        for col_id in range(1, worksheet.max_column + 1):
            col_fill[col_id] = False
            for row_id in range(1, worksheet.max_row + 1):
                if worksheet.cell(row_id, col_id).value:
                    col_fill[col_id] = True
                    break
        last_col = worksheet.max_column
        for col_id, not_empty in sorted([(col_id, not_empty) for col_id, not_empty in col_fill.items()], reverse=True,
                                        key=lambda x: x[0]):
            if not_empty:
                if last_col == worksheet.max_column:
                    last_col += 1
                break
            last_col = col_id

        height = 0
        for row_id in range(1, last_row):
            if worksheet.row_dimensions[row_id].height:
                height += worksheet.row_dimensions[row_id].height
            else:
                height += DEFAULT_HEIGHT
        width = 0
        for col_id in range(1, last_col):
            if worksheet.column_dimensions[get_column_letter(col_id)].width:
                width += worksheet.column_dimensions[get_column_letter(col_id)].width
            else:
                width += DEFAULT_WIDTH
        if height == 0 or width == 0:
            continue

        img = cv2.imread(str((images_dir / f"{page_num}.png").absolute()))
        img_shape = img.shape[:2]

        blocks = []
        if last_row > 1000:
            blocks.append(compute_tables(worksheet, last_row, last_col))

        if not blocks:
            inf_tables, headers = cascade_rcnn_detector.inference_image(images_dir / f"{page_num}.png", padding=200)

            y_scale = img_shape[0] / height
            x_scale = img_shape[1] / width

            tables = []
            table_zones = []
            for inf_table in inf_tables:
                rows_in_table = []
                prev_coord = 0
                for row_id in range(1, last_row):
                    coord = prev_coord + (worksheet.row_dimensions[row_id].height if worksheet.row_dimensions[
                        row_id].height else DEFAULT_HEIGHT)
                    if inf_table.bbox.top_left_y < prev_coord * y_scale < inf_table.bbox.bottom_right_y \
                            and inf_table.bbox.top_left_y < coord * y_scale < inf_table.bbox.bottom_right_y:
                        rows_in_table.append((row_id, int(prev_coord * y_scale), int(coord * y_scale)))
                    elif inf_table.bbox.top_left_y < prev_coord * y_scale < inf_table.bbox.bottom_right_y \
                            and not inf_table.bbox.top_left_y < coord * y_scale < inf_table.bbox.bottom_right_y:
                        if (inf_table.bbox.bottom_right_y - prev_coord * y_scale) / (
                                coord * y_scale - prev_coord * y_scale) > 0.3:
                            rows_in_table.append((row_id, int(prev_coord * y_scale), int(coord * y_scale)))
                    elif not inf_table.bbox.top_left_y < prev_coord * y_scale < inf_table.bbox.bottom_right_y \
                            and inf_table.bbox.top_left_y < coord * y_scale < inf_table.bbox.bottom_right_y:
                        if (coord * y_scale - inf_table.bbox.top_left_y) / (
                                coord * y_scale - prev_coord * y_scale) > 0.3:
                            rows_in_table.append((row_id, int(prev_coord * y_scale), int(coord * y_scale)))
                    prev_coord = coord

                cols_in_table = []
                prev_coord = 0
                for col_id in range(1, last_col):
                    coord = prev_coord + (worksheet.column_dimensions[get_column_letter(col_id)].width
                                          if worksheet.column_dimensions[
                        get_column_letter(col_id)].width else DEFAULT_WIDTH)
                    if inf_table.bbox.top_left_x < prev_coord * x_scale < inf_table.bbox.bottom_right_x \
                            and inf_table.bbox.top_left_x < coord * x_scale < inf_table.bbox.bottom_right_x:
                        cols_in_table.append((col_id, int(prev_coord * x_scale), int(coord * x_scale)))
                    elif inf_table.bbox.top_left_x < prev_coord * x_scale < inf_table.bbox.bottom_right_x \
                            and not inf_table.bbox.top_left_x < coord * x_scale < inf_table.bbox.bottom_right_x:
                        if (inf_table.bbox.bottom_right_x - prev_coord * x_scale) / (
                                coord * x_scale - prev_coord * x_scale) > 0.3:
                            cols_in_table.append((col_id, int(prev_coord * x_scale), int(coord * x_scale)))
                    elif not inf_table.bbox.top_left_x < prev_coord * x_scale < inf_table.bbox.bottom_right_x \
                            and inf_table.bbox.top_left_x < coord * x_scale < inf_table.bbox.bottom_right_x:
                        if (coord * x_scale - inf_table.bbox.top_left_x) / (
                                coord * x_scale - prev_coord * x_scale) > 0.3:
                            cols_in_table.append((col_id, int(prev_coord * x_scale), int(coord * x_scale)))
                    prev_coord = coord

                if not rows_in_table or not cols_in_table:
                    continue

                rows = [row for row, start, end in rows_in_table]
                cols = [col for col, start, end in cols_in_table]

                m_ranges = []

                for m_range in worksheet.merged_cells.ranges:
                    if (m_range.min_row in rows or m_range.max_row in rows) \
                            and (m_range.min_col in cols or m_range.max_col in cols):
                        rows = list(range(min(m_range.min_row, min(rows)), max(m_range.max_row, max(rows)) + 1))
                        cols = list(range(min(m_range.min_col, min(cols)), max(m_range.max_col, max(cols)) + 1))
                        m_ranges.append(m_range)

                rows_in_table = []
                prev_coord = 0
                for row_id in range(1, last_row):
                    coord = prev_coord + (worksheet.row_dimensions[row_id].height
                                          if worksheet.row_dimensions[row_id].height else DEFAULT_HEIGHT)
                    if row_id in rows:
                        rows_in_table.append((row_id, int(prev_coord * y_scale), int(coord * y_scale)))
                    prev_coord = coord

                cols_in_table = []
                prev_coord = 0
                for col_id in range(1, last_col):
                    coord = prev_coord + (worksheet.column_dimensions[get_column_letter(col_id)].width
                                          if worksheet.column_dimensions[
                        get_column_letter(col_id)].width else DEFAULT_WIDTH)
                    if col_id in cols:
                        cols_in_table.append((col_id, int(prev_coord * x_scale), int(coord * x_scale)))
                    prev_coord = coord

                cells: List[CellLinked] = []
                for t_col, (col, c_start, c_end) in enumerate(cols_in_table):
                    for t_row, (row, r_start, r_end) in enumerate(rows_in_table):
                        in_merged = False
                        for m_range in m_ranges:
                            if (row, col) in list(m_range.cells):
                                in_merged = True
                                if col == m_range.start_cell.col_idx and row == m_range.start_cell.row:
                                    cells.append(
                                        CellLinked(
                                            top_left_y=r_start,
                                            top_left_x=c_start,
                                            bottom_right_y=r_end,
                                            bottom_right_x=c_end,
                                            col=t_col,
                                            row=t_row,
                                            col_span=m_range.max_col - m_range.min_col + 1,
                                            row_span=m_range.max_row - m_range.min_row + 1,
                                            text_boxes=[
                                                TextField(
                                                    bbox=BorderBox(
                                                        top_left_y=r_start,
                                                        top_left_x=c_start,
                                                        bottom_right_y=
                                                        [r_e for r, r_s, r_e in rows_in_table if r == m_range.max_row][
                                                            0] if [r_e for r, r_s, r_e in rows_in_table if
                                                                   r == m_range.max_row] else rows_in_table[-1][2],
                                                        bottom_right_x=
                                                        [c_e for c, c_s, c_e in cols_in_table if c == m_range.max_col][
                                                            0] if [c_e for c, c_s, c_e in cols_in_table if
                                                                   c == m_range.max_col] else cols_in_table[-1][2]
                                                    ),
                                                    text=extract_cell_value(m_range.start_cell)
                                                )
                                            ]
                                        )
                                    )
                        if not in_merged:
                            cells.append(
                                CellLinked(
                                    top_left_y=r_start,
                                    top_left_x=c_start,
                                    bottom_right_y=r_end,
                                    bottom_right_x=c_end,
                                    col=t_col,
                                    row=t_row,
                                    col_span=1,
                                    row_span=1,
                                    text_boxes=[
                                        TextField(
                                            bbox=BorderBox(
                                                top_left_y=r_start,
                                                top_left_x=c_start,
                                                bottom_right_y=r_end,
                                                bottom_right_x=c_end
                                            ),
                                            text=extract_cell_value(worksheet.cell(row, col))
                                        )
                                    ]
                                )
                            )
                struct_table = StructuredTable(
                    bbox=BorderBox(
                        top_left_y=rows_in_table[0][1],
                        top_left_x=cols_in_table[0][1],
                        bottom_right_x=rows_in_table[-1][2],
                        bottom_right_y=cols_in_table[-1][2],
                    ),
                    cells=cells
                )

                struct_table_headered = get_headers_using_structured(struct_table, headers)
                tables.append(struct_table_headered)
                head_cells = []
                for pack in struct_table_headered.header:
                    head_cells.extend(pack)
                for cell in head_cells:
                    col = cols[cell.col]
                    row = rows[cell.row]
                    col_span = cell.col_span
                    row_span = cell.row_span
                    for r in range(row, row + row_span):
                        for c in range(col, col + col_span):
                            worksheet.cell(r, c).fill = HEADER_FILL

                table_zone = (rows_in_table[0][0], cols_in_table[0][0], rows_in_table[-1][0], cols_in_table[-1][0])
                table_zones.append(table_zone)

            blocks = []
            blocks.extend(tables)
            prev_row_coord = 0
            for row_id in range(1, last_row):
                row_coord = prev_row_coord + (worksheet.row_dimensions[row_id].height
                                              if worksheet.row_dimensions[row_id].height else DEFAULT_HEIGHT)
                prev_col_coord = 0
                for col_id in range(1, last_col):
                    col_coord = prev_col_coord + (worksheet.column_dimensions[get_column_letter(col_id)].width
                                                  if worksheet.column_dimensions[
                        get_column_letter(col_id)].width else DEFAULT_WIDTH)
                    if worksheet.cell(row_id, col_id).value and not any(
                            [y1 <= row_id < y2 and x1 <= col_id < x2 for y1, x1, y2, x2 in table_zones]):
                        text_field = TextField(
                            bbox=BorderBox(
                                top_left_x=prev_col_coord * x_scale,
                                top_left_y=prev_row_coord * y_scale,
                                bottom_right_x=col_coord * x_scale,
                                bottom_right_y=row_coord * y_scale
                            ),
                            text=extract_cell_value(worksheet.cell(row_id, col_id))
                        )
                        blocks.append(text_field)
                    prev_col_coord = col_coord
                prev_row_coord = row_coord

        pages.append(page_to_dict(Page(
            page_num=page_num,
            bbox=BorderBox(
                top_left_x=0,
                top_left_y=0,
                bottom_right_x=img_shape[1],
                bottom_right_y=img_shape[0]
            ),
            tables=blocks
        )))
    workbook.save(str(xlsx_path.absolute()))
    workbook.close()
    return pages


def crop_padding(inf_results: List[Tuple[List[InferenceTable], List[Cell]]], padding: int):
    for inf_tables, headers in inf_results:
        for table in inf_tables:
            table.bbox.top_left_x = table.bbox.top_left_x - padding
            table.bbox.top_left_y = table.bbox.top_left_y - padding
            table.bbox.bottom_right_x = table.bbox.bottom_right_x - padding
            table.bbox.bottom_right_y = table.bbox.bottom_right_y - padding
            for cell in table.tags:
                cell.top_left_x = cell.top_left_x - padding
                cell.top_left_y = cell.top_left_y - padding
                cell.bottom_right_x = cell.bottom_right_x - padding
                cell.bottom_right_y = cell.bottom_right_y - padding
        for header in headers:
            header.top_left_x = header.top_left_x - padding
            header.top_left_y = header.top_left_y - padding
            header.bottom_right_x = header.bottom_right_x - padding
            header.bottom_right_y = header.bottom_right_y - padding


def run_excel_job(file: str, outpath: str):
    out_dir = Path(outpath) / Path(file).name
    pdf_path = prepare_for_inference(Path(file), out_dir)
    images_dir = convert_pdf_to_images(pdf_path, out_dir, True)

    create_copy_xlsx(Path(file), out_dir, 'with_header.xlsx')

    pages = match_inf_res(
        out_dir / 'with_header.xlsx',
        images_dir
    )

    document = {
        'doc_name': str(Path(file).name),
        'pages': pages
    }
    save_document(document, out_dir / 'document.json')


if __name__ == "__main__":
    run_excel_job(["/Users/matvei_vargasov/code/novatris/excel_extractor/base/samples/Examples.xlsx"])
